from openpyxl import load_workbook
import os

col = ['Data Center','Task ID','Version','Pending','Ready','Queued','Running','Error','Cancel','Done','Status','%Cmpl', 'Started at datetime','Completed at datetime','Time to Complete datetime']
wb = load_workbook('read.xlsx')
sheet_obj = wb.active
rows = sheet_obj.max_row
columns = sheet_obj.max_column
spacing = [22, 8, 10, 28]
arr = []

# https://youtu.be/RvqmZLuCQrw

def fixed_lenght(text,lenght):
    if len(text) > lenght:
        text = str(text[:lenght])
    elif len(text) < lenght:
        text = (text + " " * lenght)[:lenght]
    return text

# print loop the heading
for f in range(0, len(col)):
    if f == 0:
        print(f'{fixed_lenght(str(col[f]), spacing[0])}', end = " ")
    elif f > 0 and f < 10:
        print(f'{fixed_lenght(str(col[f]), spacing[1])}', end = "")
    elif f == 10:
        print(f'{fixed_lenght(str(col[f]), spacing[2])}', end = "")
    elif f >= 12:
        print(f'{fixed_lenght(str(col[f]), spacing[3])}', end = "")
print()

# store the values in 'arr'
for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        cell_obj = sheet_obj.cell(column = j,row = i)
        if cell_obj.value == None:
            arr.append(str(" "))
        else:
            arr.append(str(cell_obj.value))

    # loop print the values
    for k in range(0, len(arr)):
        if k == 0:
            print(f'{fixed_lenght(arr[k], spacing[0])}', end = " ")
        elif k > 0 and k < 10:
            print(f'{fixed_lenght(arr[k], spacing[1])}', end = "")
        elif k == 10:
            print(f'{fixed_lenght(arr[k], spacing[2])}', end = " ")
        elif k >= 12:
            print(f'{fixed_lenght(arr[k], spacing[3])}', end = "")

    # reset array as we are going to use it multiple times
    arr = []
    print()

os.system('pause')